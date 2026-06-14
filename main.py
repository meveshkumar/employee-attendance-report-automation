import pandas as pd
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import smtplib
from email.message import EmailMessage

# Load attendance data
df = pd.read_excel("attendance.xlsx")

today = date.today()

report_file = f"Attendance_Report_{today}.xlsx"

# Office start time
office_time = datetime.strptime("09:15", "%H:%M")

# Function to determine attendance status
def check_status(login_time):
    try:
        login = datetime.strptime(str(login_time), "%H:%M")

        if login <= office_time:
            return "On Time"
        else:
            return "Late"

    except:
        return "Absent"

# Create Status column
df["Status"] = df["Login Time"].apply(check_status)

# Summary calculations
total_employees = len(df)
on_time = len(df[df["Status"] == "On Time"])
late = len(df[df["Status"] == "Late"])
absent = len(df[df["Status"] == "Absent"])

# Create summary dataframe
summary_df = pd.DataFrame({
    "Metric": [
        "Total Employees",
        "On Time",
        "Late",
        "Absent"
    ],
    "Count": [
        total_employees,
        on_time,
        late,
        absent
    ]
})
# Department-wise analytics
department_summary = pd.crosstab(
    df["Department"],
    df["Status"]
)

department_summary["Total Employees"] = department_summary.sum(axis=1)

department_summary = department_summary.reset_index()

# Write to Excel with multiple sheets
with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Attendance", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    department_summary.to_excel(
        writer,
        sheet_name="Department Analysis",
        index=False
    )

# Open workbook for formatting
wb = load_workbook(report_file)

attendance_sheet = wb["Attendance"]

# Make headers bold
for cell in attendance_sheet[1]:
    cell.font = Font(bold=True)

# Auto-adjust column widths
for column in attendance_sheet.columns:

    max_length = 0

    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    adjusted_width = max_length + 2

    attendance_sheet.column_dimensions[
        column[0].column_letter
    ].width = adjusted_width

# Create colors
yellow_fill = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FF9999",
    end_color="FF9999",
    fill_type="solid"
)

# Highlight rows
for row in attendance_sheet.iter_rows(min_row=2):

    status = row[4].value

    if status == "Late":
        for cell in row:
            cell.fill = yellow_fill

    elif status == "Absent":
        for cell in row:
            cell.fill = red_fill

# Save workbook
wb.save(report_file)

print("\n===== ATTENDANCE SUMMARY =====")
print(f"Total Employees : {total_employees}")
print(f"On Time         : {on_time}")
print(f"Late            : {late}")
print(f"Absent          : {absent}")

print("\n===== DEPARTMENT ANALYSIS =====")
print(department_summary)

print(f"\nReport File: {report_file}")

print("\nAttendance report generated successfully!")

# Email settings
EMAIL_ADDRESS = "your_email@gmail.com"
APP_PASSWORD = "your_app_password"

msg = EmailMessage()

msg["Subject"] = "Daily Attendance Report"
msg["From"] = EMAIL_ADDRESS
msg["To"] = EMAIL_ADDRESS

msg.set_content(
    "Please find attached the latest attendance report."
)

# Attach report
with open(report_file, "rb") as file:
    msg.add_attachment(
        file.read(),
        maintype="application",
        subtype="octet-stream",
        filename=report_file
    )

# Send email
with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
    smtp.login(EMAIL_ADDRESS, APP_PASSWORD)
    smtp.send_message(msg)

print("Email sent successfully!")